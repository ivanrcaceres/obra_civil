# -*- coding: utf-8 -*-
import io
from odoo import models, fields, api, exceptions, _
from odoo.http import request

from odoo.exceptions import ValidationError
from odoo import fields, models, exceptions, api
from datetime import datetime, timedelta, time, date
# -*- coding: utf-8 -*-
from odoo import http
# import certificados
# from django.http.response import JsonResponse, HttpResponse
from string import ascii_uppercase

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Alignment, Side, PatternFill
from openpyxl import Workbook

import xlwt
import base64
import xlsxwriter
from io import StringIO



from odoo.addons.web.controllers.main import serialize_exception,content_disposition

class Certificados2(http.Controller):
    @http.route('/certificados/certificados/<string:palabra>', auth='public')
    def index(self,palabra=None, **kw):
        # return 'holis'

        print ('palabra')
        print (palabra)

        x = palabra.split('-')
        id_proyecto = x[0]
        vector_avances = []
        n_avenaces = len(x)
        for i in range(1, n_avenaces):
            vector_avances.append(int(x[i]))

        ultimos_avances = request.env['certificados.avance'].browse(vector_avances)
        print (ultimos_avances)
        todos_los_avances = request.env['certificados.avance'].search([('project_id', '=', int(id_proyecto))])
        print (todos_los_avances)

        servicios_del_proyecto = request.env['certificados.proyecto_detalle_servicio'].search([('project_id', '=', int(id_proyecto))])

        print (servicios_del_proyecto)

        suma1 = []
        for a1 in servicios_del_proyecto:
            suma = 0
            for a2 in todos_los_avances:
                for a3 in a2.avance_detalle_ids:
                    if a1.product_id.id == a3.product_id.id:
                        print ('jo')
                        suma  = suma + int(a3.quantity)
                        print (suma)
            suma1.append(suma)

        print (servicios_del_proyecto)
        print (suma1)

        suma12 = []
        for a1 in servicios_del_proyecto:
            suma = 0
            for a2 in ultimos_avances:
                for a3 in a2.avance_detalle_ids:
                    if a1.product_id.id == a3.product_id.id:
                        print ('jo')
                        suma = suma + int(a3.quantity)
                        print (suma)
            suma12.append(suma)

        print (servicios_del_proyecto)
        print (suma12)

        suma3 = []
        aux10 = len(suma1)
        for i4 in range(aux10):
            suma3.append(suma1[i4] - suma12[i4])

        # suma1 = vecto que tiene todas las sumas de todos los avances
        # suma12 = vecto que tiene todas las sumas del ultimo avance
        # suma3 = vector que tiene todas las sumas de todos menos los ultimos avances

        print (suma3)









        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'Item')
        worksheet.write(0, 1, 'Codigo Catalogo')
        worksheet.write(0, 2, 'Descripcion del bien')
        worksheet.write(0, 3, 'Unidad de medida')
        worksheet.write(0, 4, 'Presentacion')
        worksheet.write(0, 5, 'Cantidad')
        worksheet.write(0, 6, 'Precio Unitario (iva incluido)')
        worksheet.write(0, 7, 'Precio Total (iva incluido)')

        worksheet.write(0, 8, 'Acumulado Anterior')
        worksheet.write(0, 9, 'Presente Certificado')
        worksheet.write(0, 10, 'Acumulado Total')

        worksheet.write(0, 11, 'Acumulado Anterior')
        worksheet.write(0, 12, 'Presente Certificado')
        worksheet.write(0, 13, 'Acumulado Total')




        c = 1
        for serviciox in servicios_del_proyecto:
            if str(serviciox.product_id.default_code) != 'False':
                worksheet.write(c, 1, str(serviciox.product_id.default_code))
            worksheet.write(c, 2, str(serviciox.product_id.name))
            worksheet.write(c, 3, str(serviciox.unidad_de_medida.name))
            worksheet.write(c, 4, str('Unidad'))
            worksheet.write(c, 5, serviciox.quantity)
            worksheet.write(c, 6, serviciox.price_unit)
            worksheet.write(c, 7, int(serviciox.price_unit)*int(serviciox.quantity))
            c = c + 1

        vv = []
        for serviciox in servicios_del_proyecto:
            vv.append(serviciox.price_unit)

        c=1
        for s in suma3:
            worksheet.write(c, 8, s)
            c = c + 1

        c = 1
        for s in suma12:
            worksheet.write(c, 9, s)
            c = c + 1

        c = 1
        for s in suma1:
            worksheet.write(c, 10, s)
            c = c + 1

        cn = len(vv)
        c = 1
        for a41 in range(cn):
            worksheet.write(c, 11, int(vv[a41]*suma3[a41]))
            c = c + 1

        cn = len(vv)
        c = 1
        for a41 in range(cn):
            worksheet.write(c, 12, int(vv[a41] * suma12[a41]))
            c = c + 1

        cn = len(vv)
        c = 1
        for a41 in range(cn):
            worksheet.write(c, 13, int(vv[a41] * suma1[a41]))
            c = c + 1


        workbook.close()
        output.seek(0)
        print (self)

        return request.make_response(output,
                                     [('Content-Type',
                                       'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet'),
                                      ('Content-Disposition', content_disposition('libro_diario.xlsx'))])


class Certificados3(http.Controller):
    @http.route('/certificados/clientes/<string:palabra>', auth='public')
    def index(self,palabra=None, **kw):
        print ('palabra2')
        print (palabra)

        x = palabra.split('-')
        id_contrato = x[0]
        vector_avances = []
        n_avenaces = len(x)
        for i in range(1, n_avenaces):
            vector_avances.append(int(x[i]))

        ultimos_avances = request.env['certificados.avance'].browse(vector_avances)
        print('ultimos_avances')
        print(ultimos_avances)
        # todos_los_avances = request.env['certificados.avance'].search([('project_id', '=', int(id_proyecto))])
        # print (todos_los_avances)

        proyectos_del_contrato = request.env['certificados.detalle_proyectos'].search([('contrato_cliente_id', '=', int(id_contrato))])
        print(proyectos_del_contrato)

        detalles_del_contrato = request.env['certificados.contrato_cliente_detalle'].search([('contrato_cliente_id', '=', int(id_contrato))])
        print(detalles_del_contrato)

        servicios = []
        for a1 in detalles_del_contrato:
            servicios.append(a1.product_id)
        print('servicios:')
        print(servicios)

        vector_objeto_servicio = []
        tamanodevector = len(servicios)
        for a2 in range(0,tamanodevector):
            c = 0
            for a3 in range(0, tamanodevector):
                if servicios[a2].id == servicios[a3].id:
                    c = c + 1
            if c < 2:
                vector_objeto_servicio.append(servicios[a2])
        print('los servicios en un vecto')
        print(vector_objeto_servicio)

        # aqui recorro todos los proyectos
        vector_objecto_avances = []
        for proyecto01 in proyectos_del_contrato:
            avance01 = request.env['certificados.avance'].browse([int(proyecto01.proyecto_id.id)])
            for avance_p01 in avance01:
                if avance_p01 != '':
                    vector_objecto_avances.append(avance_p01)

        print('los avances de todos los proyectos')
        print(vector_avances)

        todoslosavancesparaelcalculo = request.env['certificados.avance'].browse(vector_avances)
        print('todoslosavancesparaelcalculo')
        print(todoslosavancesparaelcalculo)

        vector_aux_para_ids_servicios = []
        for servicio001 in vector_objeto_servicio:
            vector_aux_para_ids_servicios.append(servicio001.id)
        print('vector ids servicios')
        print(vector_aux_para_ids_servicios)

        servicios_de_contrato = request.env['product.product'].browse(vector_aux_para_ids_servicios)
        print('todos los servicios del contrato')
        print(servicios_de_contrato)

        ## aqui abajo se hacen los calculos del acumulado de todos los servicios que se hizo en el contrato,
        ## de todos los proyectop

        suma1 = []
        for a1 in servicios_de_contrato:
            suma = 0
            for a2 in todoslosavancesparaelcalculo:
                for a3 in a2.avance_detalle_ids:
                    if a1.id == a3.product_id.id:
                        print('jo')
                        suma = suma + int(a3.quantity)
                        print(suma)
            suma1.append(suma)

        print('suma1: vecto que tiene todas las sumas de todos los avances')
        print(suma1)

        suma12 = []
        for a1 in servicios_de_contrato:
            suma = 0
            for a2 in ultimos_avances:
                for a3 in a2.avance_detalle_ids:
                    if a1.id == a3.product_id.id:
                        print('jo')
                        suma = suma + int(a3.quantity)
                        print(suma)
            suma12.append(suma)

        print('suma12: vecto que tiene todas las sumas del ultimo avance')
        print(suma12)

        suma3 = []
        aux10 = len(suma1)
        for i4 in range(aux10):
            suma3.append(suma1[i4] - suma12[i4])

        # suma1 = vecto que tiene todas las sumas de todos los avances
        # suma12 = vecto que tiene todas las sumas del ultimo avance
        # suma3 = vector que tiene todas las sumas de todos menos los ultimos avances

        print('suma3 : vector que tiene todas las sumas de todos menos los ultimos avances')
        print(suma3)


        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet()
        worksheet.write(0, 0, 'Item')
        worksheet.write(0, 1, 'Codigo Catalogo')
        worksheet.write(0, 2, 'Descripcion del bien')
        worksheet.write(0, 3, 'Unidad de medida')
        worksheet.write(0, 4, 'Presentacion')
        worksheet.write(0, 5, 'Cantidad')
        worksheet.write(0, 6, 'Precio Unitario (iva incluido)')
        worksheet.write(0, 7, 'Precio Total (iva incluido)')

        worksheet.write(0, 8, 'Acumulado Anterior')
        worksheet.write(0, 9, 'Presente Certificado')
        worksheet.write(0, 10, 'Acumulado Total')

        worksheet.write(0, 11, 'Acumulado Anterior')
        worksheet.write(0, 12, 'Presente Certificado')
        worksheet.write(0, 13, 'Acumulado Total')


        # aqui abajo se cargan los calculos al excel


        workbook.close()
        output.seek(0)


        return request.make_response(output,
                                     [('Content-Type',
                                       'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet'),
                                      ('Content-Disposition', content_disposition('libro_diario.xlsx'))])